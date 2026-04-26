"""
2026년 5월 야채/과일 가격 추론기

flow-price-forecast/input/ 의 과거 흐름표 (2023, 2024, 2025) + flow-price-forecast/input/recent/ 의 최근
서울청과(주) 가격을 결합하여 2026년 5월 가격을 추론하고
flow-price-forecast/output/야채,과일 흐름표(2026)_predicted.xlsx 를 생성합니다.

추론 모델 (단순 비율 보정):
    base[일자]      = 2025년 5월 같은 일자 가격 (없으면 2024 → 2023 폴백)
    recent_ratio    = 최근 1주일(서울청과) 평균가 / 2025년 4월말~5월초 가격
    예측[일자]      = base[일자] × recent_ratio

recent_ratio 산출이 불가능한 품목은 base 그대로(=2025 5월) 사용합니다.

사용법:
    python flow-price-forecast/scripts/predict_may_prices.py
    python flow-price-forecast/scripts/predict_may_prices.py --recent path/to/recent.csv
    python flow-price-forecast/scripts/predict_may_prices.py --target-year 2026 --target-month 5
"""

from __future__ import annotations

import argparse
import csv
import io
import re
import statistics
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
INPUT_DIR    = SCRIPT_DIR.parent / "input"
RECENT_DIR   = INPUT_DIR / "recent"
OUTPUT_DIR   = SCRIPT_DIR.parent / "output"

DATE_RE = re.compile(r"(\d{1,2})월\s*(\d{1,2})일")


@dataclass
class PriceRow:
    """A single (item, year, month, day) → price observation."""
    item: str
    year: int
    month: int
    day: int
    price: float


@dataclass
class FlowSheet:
    """One month-sheet's parsed structure: column groups + rows.

    Each group is (label_col_idx, [date_col_idx, ...], [(month, day), ...]).
    Rows are kept as the original openpyxl values so that we can re-emit a
    sheet with the same shape and just swap in predicted prices.
    """
    title: str
    column_groups: list[tuple[int, list[int], list[tuple[int, int]]]]
    raw_rows: list[list]  = field(default_factory=list)


# ---------------------------------------------------------------------------
# Parsing the historical 흐름표 workbook
# ---------------------------------------------------------------------------

def parse_flow_sheet(ws) -> FlowSheet | None:
    """Detect column groups by scanning row 1 for repeated 품목 + 날짜 patterns."""
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    groups: list[tuple[int, list[int], list[tuple[int, int]]]] = []
    i = 0
    while i < len(header):
        v = (header[i] or "")
        if isinstance(v, str) and v.strip() == "품목":
            label_col = i
            date_cols: list[int] = []
            dates: list[tuple[int, int]] = []
            j = i + 1
            while j < len(header):
                cell = header[j]
                if isinstance(cell, str) and cell.strip() == "품목":
                    break
                m = DATE_RE.search(str(cell or ""))
                if m:
                    date_cols.append(j)
                    dates.append((int(m.group(1)), int(m.group(2))))
                j += 1
            if date_cols:
                groups.append((label_col, date_cols, dates))
            i = j
        else:
            i += 1
    if not groups:
        return None

    raw_rows: list[list] = []
    for r in range(2, ws.max_row + 1):
        raw_rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])
    return FlowSheet(title=ws.title, column_groups=groups, raw_rows=raw_rows)


def load_history(input_dir: Path) -> tuple[list[PriceRow], dict[int, dict[str, FlowSheet]]]:
    """Walk `야채,과일 흐름표(YYYY).xlsx` files and flatten observations.

    Returns:
        observations — flat list of every (item, year, month, day, price)
        by_year_month — {year: {month_label: FlowSheet}} for re-emitting layouts
    """
    observations: list[PriceRow] = []
    by_year_month: dict[int, dict[str, FlowSheet]] = {}

    for path in sorted(input_dir.glob("*흐름표*.xlsx")):
        m = re.search(r"\((\d{4})\)", path.name)
        if not m:
            continue
        year = int(m.group(1))
        wb = openpyxl.load_workbook(path, data_only=True)
        by_year_month[year] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet = parse_flow_sheet(ws)
            if not sheet:
                continue
            by_year_month[year][sheet_name] = sheet
            for label_col, date_cols, dates in sheet.column_groups:
                for row in sheet.raw_rows:
                    if label_col >= len(row):
                        continue
                    name = row[label_col]
                    if not name or not isinstance(name, str):
                        continue
                    name = name.strip()
                    if not name:
                        continue
                    for col, (mo, dy) in zip(date_cols, dates):
                        if col >= len(row):
                            continue
                        v = row[col]
                        if not isinstance(v, (int, float)) or v <= 0:
                            continue
                        observations.append(PriceRow(name, year, mo, dy, float(v)))
    return observations, by_year_month


# ---------------------------------------------------------------------------
# Recent-week prices (서울청과(주))
# ---------------------------------------------------------------------------

def load_recent(path: Path | None) -> dict[str, float]:
    """Load `{item: price}` from a CSV/Excel file. Empty if file is missing."""
    if path is None:
        # default: flow-price-forecast/input/recent/recent_prices.csv (if present)
        default = RECENT_DIR / "recent_prices.csv"
        if default.exists():
            path = default
        else:
            # also accept any single csv/xlsx in the recent folder
            cands = sorted([*RECENT_DIR.glob("*.csv"), *RECENT_DIR.glob("*.xlsx")])
            if cands:
                path = cands[0]
            else:
                return {}

    if not path.exists():
        print(f"[recent] {path} 없음 — 폴백 모델만 사용")
        return {}

    out: dict[str, float] = {}
    if path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get("품목") or row.get("item") or "").strip()
                price = row.get("가격") or row.get("평균가") or row.get("price")
                if not name or price in (None, ""):
                    continue
                try:
                    out[name] = float(str(price).replace(",", ""))
                except ValueError:
                    pass
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [str(c.value or "").strip() for c in ws[1]]
        try:
            ni = next(i for i, h in enumerate(header) if h in ("품목", "item", "품목명"))
            pi = next(i for i, h in enumerate(header) if h in ("가격", "평균가", "price"))
        except StopIteration:
            print(f"[recent] {path} 헤더에서 품목/가격 컬럼을 찾지 못함")
            return {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if ni >= len(row) or pi >= len(row):
                continue
            name = str(row[ni] or "").strip()
            v = row[pi]
            if not name or v in (None, ""):
                continue
            try:
                out[name] = float(str(v).replace(",", ""))
            except ValueError:
                pass

    print(f"[recent] {path.name} → {len(out)}개 품목")
    return out


# ---------------------------------------------------------------------------
# Prediction
# ---------------------------------------------------------------------------

def latest_price(observations: list[PriceRow], item: str, year: int, month: int) -> float | None:
    """Get a representative price for (item, year, month) — mean of all dates."""
    vals = [o.price for o in observations
            if o.item == item and o.year == year and o.month == month]
    if not vals:
        return None
    return statistics.mean(vals)


def template_for(observations: list[PriceRow], item: str, target_month: int,
                 prefer_year: int) -> dict[tuple[int, int], float] | None:
    """Return {(month, day): price} from the most recent year that has data."""
    for y in (prefer_year, prefer_year - 1, prefer_year - 2):
        rows = [o for o in observations
                if o.item == item and o.year == y and o.month == target_month]
        if rows:
            return {(o.month, o.day): o.price for o in rows}
    return None


def predict(observations: list[PriceRow], recent: dict[str, float],
            target_year: int, target_month: int) -> dict[str, dict[tuple[int, int], float]]:
    """Predict prices for `target_year/target_month` per item, per date.

    Anchor year for the recent-ratio is `target_year - 1`. The denominator is
    the mean of (target_month - 1) prices in the anchor year — i.e., the same
    "1주일 전" relative position. Falls back to anchor-month mean if needed.
    """
    items = sorted({o.item for o in observations})
    anchor_year = target_year - 1
    anchor_prev_month = target_month - 1 if target_month > 1 else 12
    anchor_prev_year = anchor_year if target_month > 1 else anchor_year - 1

    out: dict[str, dict[tuple[int, int], float]] = {}
    for item in items:
        template = template_for(observations, item, target_month, anchor_year)
        if not template:
            continue

        ratio = 1.0
        recent_price = recent.get(item)
        if recent_price is not None:
            anchor = (latest_price(observations, item, anchor_prev_year, anchor_prev_month)
                      or latest_price(observations, item, anchor_year, target_month))
            if anchor and anchor > 0:
                ratio = recent_price / anchor

        out[item] = {date: round(price * ratio) for date, price in template.items()}
    return out


# ---------------------------------------------------------------------------
# Output writer
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL    = PatternFill("solid", fgColor="EAF1F8")
CENTER      = Alignment(horizontal="center", vertical="center")


def write_predictions(by_year_month: dict[int, dict[str, FlowSheet]],
                      predictions: dict[str, dict[tuple[int, int], float]],
                      target_year: int, target_month: int,
                      out_path: Path) -> None:
    """Reuse the previous year's sheet layout as the template, swap prices."""
    anchor = target_year - 1
    sheet_label = f"{target_month}월"
    template_sheet = by_year_month.get(anchor, {}).get(sheet_label)
    if not template_sheet:
        # fallback: any year's matching sheet
        for y in sorted(by_year_month.keys(), reverse=True):
            if sheet_label in by_year_month[y]:
                template_sheet = by_year_month[y][sheet_label]
                break
    if not template_sheet:
        raise RuntimeError(f"{sheet_label} 시트 템플릿을 어느 연도에서도 찾지 못함")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{sheet_label}_예측"

    n_cols = max((len(r) for r in template_sheet.raw_rows), default=0)
    headers: list = ["품목"] + [None] * (n_cols - 1)
    for label_col, date_cols, dates in template_sheet.column_groups:
        if label_col < n_cols:
            headers[label_col] = "품목"
        for col, (mo, dy) in zip(date_cols, dates):
            if col < n_cols:
                # Re-label dates to show the target year context (kept as text)
                headers[col] = f"{mo}월{dy}일(예측)"
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for ri, row in enumerate(template_sheet.raw_rows, start=2):
        out_row: list = list(row) + [None] * (n_cols - len(row))
        for label_col, date_cols, dates in template_sheet.column_groups:
            if label_col >= len(row):
                continue
            name = row[label_col]
            if not name or not isinstance(name, str):
                # preserve blank/non-string rows as-is
                continue
            name = name.strip()
            preds = predictions.get(name, {})
            for col, (mo, dy) in zip(date_cols, dates):
                if col < len(out_row) and (mo, dy) in preds:
                    out_row[col] = preds[(mo, dy)]
                elif col < len(out_row):
                    # leave blank when we have no prediction (rather than copying
                    # the historical value, which would be misleading)
                    out_row[col] = None
        for c, v in enumerate(out_row, start=1):
            cell = ws.cell(row=ri, column=c, value=v)
            if ri % 2 == 0:
                cell.fill = ALT_FILL
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    for col in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[ok] {out_path} ({len(predictions)}개 품목, {sheet_label} 예측)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="2026년 5월 가격 추론기")
    ap.add_argument("--input", type=Path, default=INPUT_DIR,
                    help="흐름표 xlsx가 있는 폴더 (기본: flow-price-forecast/input)")
    ap.add_argument("--recent", type=Path, default=None,
                    help="최근 1주일 서울청과 가격 CSV/XLSX (기본: flow-price-forecast/input/recent/recent_prices.csv)")
    ap.add_argument("--output", type=Path, default=None,
                    help="출력 xlsx 경로 (기본: flow-price-forecast/output/야채,과일 흐름표(YYYY)_predicted.xlsx)")
    ap.add_argument("--target-year", type=int, default=2026)
    ap.add_argument("--target-month", type=int, default=5)
    args = ap.parse_args()

    if not args.input.exists():
        print(f"입력 폴더 없음: {args.input}")
        sys.exit(1)

    print(f"[load] {args.input} 에서 흐름표 로드 중...")
    observations, by_year_month = load_history(args.input)
    if not observations:
        print("흐름표 데이터를 찾지 못함 (파일명 패턴: '*흐름표*.xlsx', 연도는 괄호 안)")
        sys.exit(1)
    years = sorted(by_year_month.keys())
    print(f"  → 관측 {len(observations)}건, 연도 {years}")

    recent = load_recent(args.recent)

    predictions = predict(observations, recent, args.target_year, args.target_month)
    print(f"[predict] {args.target_year}년 {args.target_month}월: {len(predictions)}개 품목")

    out_path = args.output or (
        OUTPUT_DIR / f"야채,과일 흐름표({args.target_year})_predicted.xlsx"
    )
    write_predictions(by_year_month, predictions,
                      args.target_year, args.target_month, out_path)


if __name__ == "__main__":
    main()
