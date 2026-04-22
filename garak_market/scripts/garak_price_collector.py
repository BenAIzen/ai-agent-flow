"""
가락시장 가격 수집기 (가락도매시장 - 아스파라거스)

매일 실행하여 아스파라거스 품목의 가격 데이터를 SQLite에 누적 저장하고
Excel 리포트를 자동 생성합니다.

사용법:
    python garak_market/scripts/garak_price_collector.py          # 어제 날짜
    python garak_market/scripts/garak_price_collector.py --date 20260411
    python garak_market/scripts/garak_price_collector.py --report  # Excel만 재생성
"""

from __future__ import annotations

import argparse
import asyncio
import io
import json
import sqlite3
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 설정 ────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent          # garak_market/scripts/
PROJECT_ROOT = SCRIPT_DIR.parent.parent                 # green_food/
DATA_DIR     = SCRIPT_DIR.parent / "data"               # garak_market/data/
DB_PATH      = DATA_DIR / "garak_prices.db"
REPORT_PATH  = DATA_DIR / "가락시장_아스파라거스_시세.xlsx"

# ── 가락시장 API 설정 ─────────────────────────────────────────────────────────
GARAK_MAIN_URL  = "https://www.garak.co.kr/youtong/G1000398/dashboard/typePrice.do"
BIX5_SHARE_URL  = "https://db.garak.co.kr:9443/shares/4b29ed90716b37d7b99eace4d29583e9"
BIX5_DS_URL     = "https://db.garak.co.kr:9443/api/datasources/40f2c32edec68ae89c0994c0f2d8dab6"

TARGET_ITM_CD   = "25301"          # 아스파라거스 국산
TARGET_ITM_NM   = "아스파라거스"
HANDL_CLSS_CD   = "2"              # 채소류

# ── UTF-8 출력 (Windows) ──────────────────────────────────────────────────────
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


# ═══════════════════════════════════════════════════════════════════════════════
# 1. DB 초기화
# ═══════════════════════════════════════════════════════════════════════════════
def init_db() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS prices (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT NOT NULL,          -- YYYY-MM-DD
            itm_nm      TEXT,                   -- 품목명
            grade       TEXT,                   -- 등급 (상/보통/하)
            unit        TEXT,                   -- 단위 (1kg 등)
            unit_qty    REAL,                   -- 단위 수량
            min_price   INTEGER,                -- 최저가
            max_price   INTEGER,                -- 최고가
            avg_price   INTEGER,                -- 평균가
            prev_diff   INTEGER,                -- 전일비
            collected_at TEXT,                  -- 수집 시각
            UNIQUE(date, itm_nm, grade)
        )
    """)
    con.commit()
    return con


# ═══════════════════════════════════════════════════════════════════════════════
# 2. 세션 획득 (Playwright → JSESSIONID)
# ═══════════════════════════════════════════════════════════════════════════════
async def _get_session_cookie() -> str:
    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()
        await page.goto(BIX5_SHARE_URL, timeout=30_000)
        await page.wait_for_timeout(3_000)
        cookies = await context.cookies()
        await browser.close()

    for c in cookies:
        if c["name"] == "JSESSIONID":
            return c["value"]
    raise RuntimeError("JSESSIONID 쿠키를 얻지 못했습니다.")


def get_session() -> requests.Session:
    """Playwright로 JSESSIONID를 획득한 뒤 requests.Session에 주입합니다."""
    jsession = asyncio.run(_get_session_cookie())
    session = requests.Session()
    session.cookies.set("JSESSIONID", jsession, domain="db.garak.co.kr")
    session.headers.update({
        "Content-Type": "application/json",
        "Referer": "https://db.garak.co.kr:9443/",
        "User-Agent": "Mozilla/5.0",
    })
    return session


# ═══════════════════════════════════════════════════════════════════════════════
# 3. 가격 데이터 수집
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_prices(session: requests.Session, target_date: str) -> list[dict]:
    """
    target_date: 'YYYYMMDD' 형식
    반환: 아스파라거스 행 리스트
    """
    payload = {
        "startDate": target_date,
        "endDate":   target_date,
        "handlClssCd": HANDL_CLSS_CD,
    }
    url = f"{BIX5_DS_URL}?dummy={int(time.time() * 1000)}"
    resp = session.post(url, json=payload, verify=False, timeout=20)
    resp.raise_for_status()

    dataset = resp.json().get("dataset", [])
    return [
        row for row in dataset
        if TARGET_ITM_NM in str(row.get("RPTV_ITM_NM", ""))
        or TARGET_ITM_CD == str(row.get("ITM_CD", ""))
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# 4. DB 저장
# ═══════════════════════════════════════════════════════════════════════════════
def save_to_db(con: sqlite3.Connection, rows: list[dict], target_date: str) -> int:
    """저장된 새 행 수를 반환합니다. 중복은 무시합니다."""
    date_str = f"{target_date[:4]}-{target_date[4:6]}-{target_date[6:]}"
    now = datetime.now().isoformat(timespec="seconds")
    inserted = 0
    for row in rows:
        try:
            con.execute(
                """INSERT OR IGNORE INTO prices
                   (date, itm_nm, grade, unit, unit_qty,
                    min_price, max_price, avg_price, prev_diff, collected_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    date_str,
                    row.get("ITM_NM", ""),
                    row.get("G_NAME", ""),
                    row.get("UNIT", ""),
                    row.get("UNIT_QTY"),
                    row.get("MI_P"),
                    row.get("MA_P"),
                    row.get("AV_P"),
                    row.get("FLT_P"),
                    now,
                ),
            )
            if con.total_changes > 0:
                inserted += 1
        except sqlite3.IntegrityError:
            pass
    con.commit()
    return inserted


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Excel 리포트 생성
# ═══════════════════════════════════════════════════════════════════════════════
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
SUBHDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
BOLD         = Font(bold=True)
CENTER       = Alignment(horizontal="center", vertical="center")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
NUM_FMT      = "#,##0"


def _style_cell(cell, fill=None, font=None, alignment=CENTER, border=THIN, number_format=None):
    if fill:         cell.fill = fill
    if font:         cell.font = font
    if alignment:    cell.alignment = alignment
    if border:       cell.border = border
    if number_format: cell.number_format = number_format


def generate_report(con: sqlite3.Connection) -> Path:
    rows = con.execute(
        """SELECT date, itm_nm, grade, unit, unit_qty,
                  min_price, max_price, avg_price, prev_diff
           FROM prices
           ORDER BY date DESC, grade"""
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "아스파라거스 시세"

    # ── 타이틀 ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "가락시장 아스파라거스 가격 현황 (서울청과 참고)"
    title_cell.font  = Font(color="FFFFFF", bold=True, size=14)
    title_cell.fill  = HEADER_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = f"수집 기준: 가락도매시장 전체 (아스파라거스 국산) | 최종 업데이트: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub.font  = Font(italic=True, size=9, color="595959")
    sub.alignment = CENTER

    # ── 헤더 ─────────────────────────────────────────────────────────────────
    headers = ["날짜", "품목명", "등급", "단위", "단위수량",
               "최저가(원)", "최고가(원)", "평균가(원)", "전일비(원)"]
    col_widths = [13, 16, 8, 8, 9, 12, 12, 12, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        _style_cell(cell, fill=SUBHDR_FILL, font=SUBHDR_FONT)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    # ── 데이터 ────────────────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows, start=4):
        fill = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL
        values = list(row)
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            num_fmt = NUM_FMT if c_idx in (5, 6, 7, 8, 9) else None
            _style_cell(cell, fill=fill, font=Font(size=10),
                        alignment=CENTER, number_format=num_fmt)
            # 전일비 색상
            if c_idx == 9 and isinstance(val, (int, float)):
                cell.font = Font(
                    size=10,
                    color=("C00000" if val < 0 else "375623" if val > 0 else "000000"),
                    bold=(val != 0),
                )

    # ── 요약 시트 ─────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("등급별 최근 7일")
    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = "아스파라거스 등급별 최근 7일 평균가"
    t2.font = Font(bold=True, size=13, color="FFFFFF")
    t2.fill = HEADER_FILL
    t2.alignment = CENTER

    for col, h in enumerate(["날짜", "상(평균)", "상(최고)", "보통(평균)", "보통(최고)", "하(평균)", "하(최고)"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        _style_cell(cell, fill=SUBHDR_FILL, font=SUBHDR_FONT)
        ws2.column_dimensions[get_column_letter(col)].width = 13

    recent = con.execute(
        """SELECT date,
                  MAX(CASE WHEN grade='상' THEN avg_price END),
                  MAX(CASE WHEN grade='상' THEN max_price END),
                  MAX(CASE WHEN grade='보통' THEN avg_price END),
                  MAX(CASE WHEN grade='보통' THEN max_price END),
                  MAX(CASE WHEN grade='하' THEN avg_price END),
                  MAX(CASE WHEN grade='하' THEN max_price END)
           FROM prices
           GROUP BY date
           ORDER BY date DESC
           LIMIT 7"""
    ).fetchall()

    for r_idx, row in enumerate(recent, start=3):
        fill = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL
        for c_idx, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            num_fmt = NUM_FMT if c_idx > 1 else None
            _style_cell(cell, fill=fill, font=Font(size=10),
                        alignment=CENTER, number_format=num_fmt)

    wb.save(REPORT_PATH)
    return REPORT_PATH


# ═══════════════════════════════════════════════════════════════════════════════
# 6. 메인
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    parser = argparse.ArgumentParser(description="가락시장 아스파라거스 가격 수집기")
    parser.add_argument("--date",   help="수집 날짜 YYYYMMDD (기본: 어제)", default=None)
    parser.add_argument("--report", action="store_true", help="DB → Excel 리포트만 재생성")
    args = parser.parse_args()

    con = init_db()

    if args.report:
        path = generate_report(con)
        print(f"[report] {path}")
        return

    # 날짜 결정 (기본: 어제 — 당일 데이터는 장 마감 후 생성)
    if args.date:
        target = args.date
    else:
        target = (date.today() - timedelta(days=1)).strftime("%Y%m%d")

    date_display = f"{target[:4]}-{target[4:6]}-{target[6:]}"
    print(f"[collect] {date_display} 아스파라거스 가격 수집 시작...")

    print("  → 세션 초기화 중...")
    session = get_session()

    print("  → API 조회 중...")
    rows = fetch_prices(session, target)

    if not rows:
        print(f"  → {date_display}: 데이터 없음 (휴장일이거나 아직 미반영)")
        return

    inserted = save_to_db(con, rows, target)
    print(f"  → {len(rows)}행 수집, {inserted}행 신규 저장")

    for row in rows:
        print(f"     [{row.get('G_NAME')}] 평균 {row.get('AV_P'):,}원 "
              f"(최저 {row.get('MI_P'):,} / 최고 {row.get('MA_P'):,}) "
              f"전일비 {row.get('FLT_P'):+,}")

    report = generate_report(con)
    print(f"  → 리포트 저장: {report}")


if __name__ == "__main__":
    main()
