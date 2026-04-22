"""
Purchase-order converter.

Reads files in purchase_order/input and writes a normalized workbook to
purchase_order/output matching the 판매일(월)보 sample layout.

Supported inputs:
  - .xlsx / .xls  (per-vendor parsers + generic fallback)
  - .html         (Birchstreet-style PO tables + generic table fallback)
  - .pdf          (pdfplumber text + table heuristics)
  - image / scanned-PDF  (handled out-of-band: see --json mode)

Image files (.jpg/.jpeg/.png) and scanned PDFs are not readable from a
pure-Python pipeline. For those, Claude Code reads the file with its built-in
vision, extracts line items as JSON, and calls this script with --json
to write the output row by row.

Output filename: <stem>_output.xlsx
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from datetime import date
from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

# Force stdout to UTF-8 for Korean on Windows cp1252 terminals.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = Path(__file__).resolve().parent        # purchase_order/scripts/
PROJECT_ROOT = SCRIPT_DIR.parent.parent              # green_food/
DEFAULT_INPUT = PROJECT_ROOT / "purchase_order" / "input"
DEFAULT_OUTPUT = PROJECT_ROOT / "purchase_order" / "output"

HEADERS = [
    "No", "일자", "품목코드", "품목명", "규격", "단위", "납기일",
    "수량", "단가", "공급가액", "부가세", "합계금액", "할인율",
    "완료여부", "마감여부", "과세구분", "문서번호", "거래처코드",
    "거래처명", "유형", "부서", "사원", "관리항목", "창고코드",
    "창고명", "프로젝트코드", "프로젝트", "품목비고",
]

SKIP_FILES = {"발주 샘플 리스트.xlsx"}

IMAGE_EXTS = {".jpg", ".jpeg", ".png"}

# Unit tokens that appear in purchase-order line items.
UNIT_TOKENS = {
    "KG", "Kg", "kg", "KGR", "G", "EA", "ea",
    "BOX", "Box", "box", "BX", "Bx",
    "PK", "pk", "PKG", "PACK", "PAC", "pack", "PC", "PCS",
    "UOM", "UNIT",
    "단", "통", "개", "팩", "봉", "묶음", "대", "마리", "병", "캔",
}


# ============================================================================ #
# Data model                                                                    #
# ============================================================================ #

@dataclass
class OrderLine:
    품목코드: str = ""
    품목명: str = ""
    규격: str = ""
    단위: str = ""
    납기일: str = ""
    수량: float | str = ""
    단가: float | str = ""
    공급가액: float | str = ""
    부가세: float | str = ""
    합계금액: float | str = ""
    과세구분: str = "면세"
    비고: str = ""


@dataclass
class OrderDoc:
    거래처명: str
    일자: str
    lines: list[OrderLine] = field(default_factory=list)


# ============================================================================ #
# Helpers                                                                       #
# ============================================================================ #

def today_iso() -> str:
    return date.today().isoformat()


def as_num(v) -> float | str:
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace(",", "").replace("₩", "").strip()
    if not s:
        return ""
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return ""


def finalize_line(line: OrderLine, fallback_tax: str = "면세") -> OrderLine:
    q = as_num(line.수량)
    p = as_num(line.단가)
    supply = as_num(line.공급가액)
    if supply == "" and isinstance(q, (int, float)) and isinstance(p, (int, float)):
        supply = q * p
    line.수량 = q
    line.단가 = p
    line.공급가액 = supply
    if not line.과세구분:
        line.과세구분 = fallback_tax
    vat = as_num(line.부가세)
    if vat == "":
        vat = round(supply * 0.1) if (line.과세구분 == "과세" and isinstance(supply, (int, float))) else 0
    line.부가세 = vat
    total = as_num(line.합계금액)
    if total == "" and isinstance(supply, (int, float)) and isinstance(vat, (int, float)):
        total = supply + vat
    line.합계금액 = total
    return line


def looks_numeric(tok: str) -> bool:
    s = tok.replace(",", "").replace("₩", "").strip()
    if not s or s in ("-", "/"):
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


# ============================================================================ #
# Excel parsers                                                                 #
# ============================================================================ #

def parse_haevichi(path: Path) -> OrderDoc:
    """해비치.xlsx — headers on row 1, data from row 2."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx.get("품목명", 2)] in (None, ""):
            continue
        line = OrderLine(
            품목코드=str(row[idx.get("품목코드", 1)] or ""),
            품목명=str(row[idx.get("품목명", 2)] or ""),
            규격=str(row[idx.get("구매규격", 3)] or ""),
            단위=str(row[idx.get("품목단위", 4)] or ""),
            수량=row[idx.get("수량", 5)],
            단가=row[idx.get("단가", 6)],
            공급가액=row[idx.get("금액", 7)],
            과세구분="면세",
            비고=str(row[idx.get("비고", 9)] or "") if idx.get("비고") is not None else "",
        )
        doc.lines.append(finalize_line(line))
    return doc


def parse_haward(path: Path) -> OrderDoc:
    """하워드.xlsx — uses the last (latest) sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = wb.sheetnames[-1]
    ws = wb[sheet_name]
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[0]
        qty = row[3] if len(row) > 3 else None
        if not name or qty in (None, "", 0, "0"):
            continue
        line = OrderLine(
            품목명=str(name),
            규격=str(row[1] or ""),
            단위=str(row[2] or ""),
            수량=qty,
            단가=row[4] if len(row) > 4 else "",
            공급가액=row[5] if len(row) > 5 else "",
            과세구분="면세",
        )
        doc.lines.append(finalize_line(line))
    return doc


def parse_themeriden(path: Path) -> OrderDoc:
    """더메리든.xlsx — wide grid with multiple day-column groups."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())
    r3 = [c.value for c in ws[3]]
    groups = [i for i, v in enumerate(r3) if v and str(v).strip() == "품명"]
    for g in groups:
        for row in ws.iter_rows(min_row=4, values_only=True):
            if g >= len(row):
                continue
            name = row[g]
            unit = row[g + 1] if g + 1 < len(row) else None
            qty = row[g + 2] if g + 2 < len(row) else None
            if not name or qty in (None, "", 0):
                continue
            line = OrderLine(
                품목명=str(name).strip(),
                단위=str(unit or ""),
                수량=qty,
                과세구분="면세",
            )
            doc.lines.append(finalize_line(line))
    return doc


def parse_heyaudrey(path: Path) -> OrderDoc:
    """헤이오드리.xlsx — 효창CK_* sheets; headers r7, data r8+; skip 농산물."""
    wb = openpyxl.load_workbook(path, data_only=True)
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())
    for sheet_name in wb.sheetnames:
        if "농산물" in sheet_name:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=8, values_only=True):
            if len(row) < 7:
                continue
            _, no, pcode, name, qty, unit, note = row[:7]
            if not name or qty in (None, "", 0) or str(name).strip() in ("X", "x"):
                continue
            line = OrderLine(
                품목코드=str(pcode or ""),
                품목명=str(name).strip(),
                단위=str(unit or ""),
                수량=qty,
                비고=str(note or ""),
                과세구분="면세",
            )
            doc.lines.append(finalize_line(line))
    return doc


def parse_hoam(path: Path) -> OrderDoc:
    """호암교수회관.xls — legacy xls with summary rows mixed in."""
    if xlrd is None:
        raise RuntimeError("xlrd is required for .xls files")
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())
    for i in range(1, ws.nrows):
        row = ws.row_values(i)
        pcode = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        if not pcode or not name or "합계" in pcode:
            continue
        if not re.match(r"^[A-Za-z0-9\-]+$", pcode):
            continue
        line = OrderLine(
            품목코드=pcode,
            품목명=name,
            규격=str(row[2] or ""),
            단위=str(row[3] or ""),
            수량=row[4],
            단가=row[5],
            공급가액=row[6],
            부가세=row[7],
            합계금액=row[8],
            납기일=str(row[9] or ""),
            비고=str(row[10] or ""),
            과세구분="과세" if as_num(row[7]) not in ("", 0) else "면세",
        )
        doc.lines.append(finalize_line(line))
    return doc


def parse_generic_xlsx(path: Path) -> OrderDoc | None:
    """Best-effort: scan sheets for a header row containing 품목명/품명 + 수량."""
    wb = openpyxl.load_workbook(path, data_only=True)
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(1, min(ws.max_row, 15) + 1):
            row = [str(c.value).strip() if c.value else "" for c in ws[r]]
            name_col = next((i for i, v in enumerate(row) if v in ("품목명", "품명")), None)
            qty_col = next((i for i, v in enumerate(row) if v in ("수량", "발주수량")), None)
            if name_col is None or qty_col is None:
                continue
            unit_col = next((i for i, v in enumerate(row) if v in ("단위", "품목단위", "재고단위")), None)
            spec_col = next((i for i, v in enumerate(row) if v in ("규격", "구매규격", "재고규격")), None)
            price_col = next((i for i, v in enumerate(row) if v in ("단가",)), None)
            code_col = next((i for i, v in enumerate(row) if v in ("품목코드", "품번", "코드")), None)
            for drow in ws.iter_rows(min_row=r + 1, values_only=True):
                if qty_col >= len(drow):
                    continue
                name = drow[name_col] if name_col < len(drow) else None
                qty = drow[qty_col]
                if not name or qty in (None, "", 0):
                    continue
                line = OrderLine(
                    품목코드=str(drow[code_col] or "") if code_col is not None and code_col < len(drow) else "",
                    품목명=str(name).strip(),
                    규격=str(drow[spec_col] or "") if spec_col is not None and spec_col < len(drow) else "",
                    단위=str(drow[unit_col] or "") if unit_col is not None and unit_col < len(drow) else "",
                    수량=qty,
                    단가=drow[price_col] if price_col is not None and price_col < len(drow) else "",
                    과세구분="면세",
                )
                doc.lines.append(finalize_line(line))
            break
    return doc if doc.lines else None


# ============================================================================ #
# HTML parser                                                                   #
# ============================================================================ #

HTML_HEADER_KEYS = {
    "품목", "품명", "제품 설명", "Product Desc.", "Product Desc",
    "Description", "Item", "Item Code", "Item Name", "Item SKU",
    "품목 SKU", "품목코드",
}
HTML_QTY_KEYS = {"수량", "Qty", "QTY", "Quantity"}


def _extract_html_row(cells: list[str], header: list[str]) -> OrderLine | None:
    """Given a data row and its header, pull out the standard fields."""
    if not cells or not any(cells):
        return None
    # Skip "부서: ..." section headers etc.
    joined = " ".join(cells)
    if re.match(r"^(부서|Department)\s*[:：]", joined):
        return None
    first = cells[0].strip()
    # A data row starts either with a numeric line number or with an item code.
    if not (re.match(r"^\d+$", first) or re.match(r"^[A-Za-z0-9][\w\.\-/]{3,}$", first)):
        return None

    # Map header columns to field names by case-insensitive keyword matching.
    def find(keys: set[str]) -> int | None:
        lower_hdr = [(h or "").lower() for h in header]
        lower_keys = [k.lower() for k in keys]
        for i, h in enumerate(lower_hdr):
            if h in lower_keys:
                return i
        for i, h in enumerate(lower_hdr):
            for k in lower_keys:
                if k and k in h:
                    return i
        return None

    code_i = find({"품목", "Item SKU", "품목 SKU", "Item Code"})
    desc_i = find({"제품 설명", "품명", "Product Desc.", "Product Desc", "Item Name", "Description"})
    pack_i = find({"팩/크기", "Pack/ Size", "Pack/Size", "Size"})
    qty_i = find({"수량", "Qty", "Quantity"})
    unit_i = find({"단위", "UOM", "Unit"})
    price_i = find({"단가", "Price", "Unit Price", "가격"})
    total_i = find({"금액", "Amount", "총계", "확장", "Extension"})

    def get(i: int | None) -> str:
        if i is None or i >= len(cells):
            return ""
        return str(cells[i] or "").strip()

    # Birchstreet data rows often have 1 more cell than the header (tax code
    # like "EKOR0" sitting between 세금 and 금액). When that happens, the 금액
    # position in data rows is shifted by +1 relative to header position.
    offset = len(cells) - len(header)
    if offset > 0 and total_i is not None:
        total_i_data = total_i + offset
    else:
        total_i_data = total_i

    def get_data(i: int | None) -> str:
        if i is None or i >= len(cells):
            return ""
        return str(cells[i] or "").strip()

    name_raw = get(desc_i)
    line = OrderLine(
        품목코드=get(code_i),
        품목명=name_raw,
        규격=get(pack_i).replace("/", "").strip(),
        단위=get(unit_i),
        수량=get(qty_i),
        단가=get(price_i),
        공급가액=get_data(total_i_data),
        과세구분="면세",
    )
    if not line.품목명 or not looks_numeric(str(line.수량)):
        return None
    return finalize_line(line)


def parse_html_generic(path: Path) -> OrderDoc | None:
    """Find the items table in an HTML PO and extract line rows."""
    if BeautifulSoup is None:
        raise RuntimeError("beautifulsoup4 is required for HTML files")
    with open(path, "rb") as f:
        soup = BeautifulSoup(f.read(), "lxml")
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())

    best_rows = []
    best_header: list[str] = []
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        # find a header row inside the table
        for hi, hr in enumerate(rows[:6]):
            hcells = [c.get_text(" ", strip=True) for c in hr.find_all(["td", "th"])]
            # Require: multi-cell row, each cell short enough to be a real header
            # (not a giant concatenated blob), and containing the expected keywords.
            if len(hcells) < 4:
                continue
            if any(len(c) > 60 for c in hcells):
                continue
            matched_header = sum(
                1 for c in hcells
                if any(k.lower() == c.lower() or (k and k.lower() in c.lower() and len(c) < 40)
                       for k in HTML_HEADER_KEYS)
            )
            matched_qty = any(
                any(k.lower() == c.lower() or (k and k.lower() in c.lower() and len(c) < 40)
                    for k in HTML_QTY_KEYS)
                for c in hcells
            )
            if matched_header >= 1 and matched_qty:
                data_rows = rows[hi + 1:]
                if len(data_rows) > len(best_rows):
                    best_rows = data_rows
                    best_header = hcells
                break

    if not best_rows:
        return None

    for tr in best_rows:
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])]
        if not cells:
            continue
        line = _extract_html_row(cells, best_header)
        if line:
            doc.lines.append(line)

    return doc if doc.lines else None


# ============================================================================ #
# PDF parser                                                                    #
# ============================================================================ #

PDF_TEXT_HEADER_RE = re.compile(
    r"(?:순번|No\.?)?\s*(?:품목\s*코드|Item\s*Code|Code)\s+"
    r"(?:품목\s*명|Item\s*Name|품명|Description)",
    re.IGNORECASE,
)

# Match a data line: optional leading number, then tokens ending in qty/price/amount.
# We use token scanning instead of one big regex for flexibility.

def _pdf_try_table_form(page) -> list[OrderLine]:
    """Look for a table whose header row contains 순번/품목코드/품목명/수량."""
    lines: list[OrderLine] = []
    tables = page.extract_tables() or []
    for t in tables:
        header_idx = None
        for ri, row in enumerate(t):
            joined = " ".join(str(c or "") for c in row)
            if ("순번" in joined or "No" in joined) and ("품목코드" in joined or "Item Code" in joined) and "품목명" in joined:
                header_idx = ri
                break
        if header_idx is None:
            continue
        header = [str(c or "").strip() for c in t[header_idx]]
        # Build column map
        def col(*names):
            for name in names:
                for i, h in enumerate(header):
                    if h and name in h:
                        return i
            return None
        c_no = col("순번", "No")
        c_code = col("품목코드", "Item Code")
        c_name = col("품목명", "Item Name")
        c_spec = col("구매규격", "규격", "Size")
        c_unit = col("품목단위", "단위", "Unit")
        c_qty = col("수량", "Qty", "Quantity")
        c_price = col("단가", "Price")
        c_amount = col("금액", "Amount")
        c_note = col("비고", "Remark")

        for row in t[header_idx + 1:]:
            if not row:
                continue
            def g(i):
                if i is None or i >= len(row) or row[i] is None:
                    return ""
                return str(row[i]).strip()
            name = g(c_name)
            qty = g(c_qty)
            if not name or not looks_numeric(qty):
                continue
            line = OrderLine(
                품목코드=g(c_code),
                품목명=name,
                규격=g(c_spec),
                단위=g(c_unit),
                수량=qty,
                단가=g(c_price),
                공급가액=g(c_amount),
                비고=g(c_note),
                과세구분="면세",
            )
            lines.append(finalize_line(line))
        if lines:
            break
    return lines


def _pdf_try_text_form(page) -> list[OrderLine]:
    """Parse item lines from extracted text using a right-anchored heuristic.

    An item line typically ends with: [unit] qty unit_price amount [maybe type]
    where qty/unit_price/amount are numeric. No "start" gate — every line is
    tested against a shape pattern, so lines that don't match are just skipped.
    """
    text = page.extract_text() or ""
    if not text:
        return []
    lines: list[OrderLine] = []
    line_no_pat = re.compile(r"^\d+\b")
    for raw in text.split("\n"):
        ln = raw.strip()
        if not ln:
            continue
        # Stop completely at footer summary lines.
        if re.match(r"^(합\s*계|Sub\s*Total|Total\s|Terms|이용\s*약관|---)", ln):
            break
        tokens = ln.split()
        if len(tokens) < 5:
            continue
        # The line must start with a recognizable item identifier:
        # - optional 1-3 digit line number, then
        # - a code token (hyphenated/dotted alphanum OR 4+ pure digits)
        start = 0
        if re.match(r"^\d{1,3}$", tokens[0]):
            start = 1
        if start >= len(tokens):
            continue
        code_tok = tokens[start]
        if not (re.match(r"^[A-Za-z0-9]+([\-\.][A-Za-z0-9]+){1,}$", code_tok)
                or re.match(r"^\d{4,}$", code_tok)):
            continue
        code = code_tok
        start += 1
        # Find the last 3 numeric tokens
        nums = [(i, t) for i, t in enumerate(tokens) if looks_numeric(t)]
        if len(nums) < 3:
            continue
        amt_i, amt_v = nums[-1]
        pri_i, pri_v = nums[-2]
        qty_i, qty_v = nums[-3]
        if amt_i - qty_i > 4:
            continue
        # qty should be within a reasonable range (guards against year numbers etc.)
        qty_float = as_num(qty_v)
        if not isinstance(qty_float, (int, float)) or qty_float <= 0 or qty_float > 10000:
            continue
        # Unit (optional) is the token right before qty
        unit = tokens[qty_i - 1] if qty_i - 1 > start else ""
        # The "name" spans from `start` up to (but not including) the unit/qty.
        # If the token right before qty looks like a unit word, exclude it; else include.
        name_end = qty_i
        if unit and unit.upper() in {u.upper() for u in UNIT_TOKENS}:
            name_end = qty_i - 1
        name_tokens = tokens[start:name_end]
        name = " ".join(name_tokens).strip()
        # Suffix after amount may contain tax-type or notes
        note = " ".join(tokens[amt_i + 1:]).strip()
        if not name:
            continue
        line = OrderLine(
            품목코드=code,
            품목명=name,
            단위=unit if unit and unit.upper() in {u.upper() for u in UNIT_TOKENS} else "",
            수량=qty_v,
            단가=pri_v,
            공급가액=amt_v,
            비고=note,
            과세구분="면세",
        )
        lines.append(finalize_line(line))
    return lines


def parse_pdf(path: Path) -> OrderDoc | None:
    if pdfplumber is None:
        raise RuntimeError("pdfplumber is required for PDF files")
    doc = OrderDoc(거래처명=path.stem, 일자=today_iso())
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                lines = _pdf_try_table_form(page)
                if not lines:
                    lines = _pdf_try_text_form(page)
                doc.lines.extend(lines)
    except Exception as e:
        print(f"[pdf error] {path.name}: {e}")
        return None
    return doc if doc.lines else None


# ============================================================================ #
# JSON-append mode (for vision-extracted rows)                                  #
# ============================================================================ #

def parse_json_doc(path: Path) -> OrderDoc:
    """Load a JSON document describing an order, written by Claude vision.

    Expected shape:
      {
        "거래처명": "네스트",
        "일자": "2026-04-12",
        "lines": [
          {"품목코드": "...", "품목명": "...", "규격": "...", "단위": "...",
           "수량": 3, "단가": 1500, "공급가액": 4500, "비고": "...",
           "과세구분": "면세"},
          ...
        ]
      }
    """
    with open(path, "rb") as f:
        data = json.loads(f.read().decode("utf-8"))
    doc = OrderDoc(
        거래처명=data.get("거래처명") or path.stem,
        일자=data.get("일자") or today_iso(),
    )
    for raw in data.get("lines", []):
        line = OrderLine(
            품목코드=str(raw.get("품목코드", "")),
            품목명=str(raw.get("품목명", "")),
            규격=str(raw.get("규격", "")),
            단위=str(raw.get("단위", "")),
            납기일=str(raw.get("납기일", "")),
            수량=raw.get("수량", ""),
            단가=raw.get("단가", ""),
            공급가액=raw.get("공급가액", ""),
            부가세=raw.get("부가세", ""),
            합계금액=raw.get("합계금액", ""),
            과세구분=str(raw.get("과세구분") or "면세"),
            비고=str(raw.get("비고", "")),
        )
        doc.lines.append(finalize_line(line))
    return doc


# ============================================================================ #
# Dispatch                                                                      #
# ============================================================================ #

EXCEL_PARSERS: dict[str, Callable[[Path], OrderDoc]] = {
    "해비치.xlsx": parse_haevichi,
    "하워드.xlsx": parse_haward,
    "더메리든.xlsx": parse_themeriden,
    "헤이오드리.xlsx": parse_heyaudrey,
    "호암교수회관.xls": parse_hoam,
}


def convert_one(path: Path, out_dir: Path) -> Path | None:
    ext = path.suffix.lower()
    if path.name in SKIP_FILES:
        print(f"[skip] {path.name} (reference file)")
        return None

    try:
        if ext in (".xlsx", ".xls"):
            parser = EXCEL_PARSERS.get(path.name)
            doc = parser(path) if parser else parse_generic_xlsx(path)
        elif ext == ".html":
            doc = parse_html_generic(path)
        elif ext == ".pdf":
            doc = parse_pdf(path)
        elif ext in IMAGE_EXTS:
            print(f"[vision] {path.name} — run via Claude vision (see SKILL.md). Skipping.")
            return None
        else:
            print(f"[skip] {path.name} (unsupported type {ext})")
            return None
    except Exception as e:
        print(f"[FAIL] {path.name}: {e}")
        return None

    if doc is None or not doc.lines:
        print(f"[skip] {path.name} (no recognizable order lines)")
        return None

    out_path = out_dir / f"{path.stem}_output.xlsx"
    write_output(doc, out_path)
    print(f"[ok]   {path.name} -> {out_path.name} ({len(doc.lines)} lines)")
    return out_path


def write_output(doc: OrderDoc, out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for i, line in enumerate(doc.lines, start=1):
        ws.append([
            i, doc.일자, line.품목코드, line.품목명, line.규격, line.단위,
            line.납기일, line.수량, line.단가, line.공급가액, line.부가세,
            line.합계금액, "", "미진행", "부", line.과세구분, "", "",
            doc.거래처명, "판매출고", "", "", "", "", "", "", "", line.비고,
        ])
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="Convert purchase-order files.")
    ap.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    ap.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    ap.add_argument("--file", type=str, default=None,
                    help="Convert only this filename from the input folder.")
    ap.add_argument("--json", type=Path, default=None,
                    help="Read a vision-extracted JSON doc and write an output xlsx. "
                         "Use --out-name to control the base name.")
    ap.add_argument("--out-name", type=str, default=None,
                    help="Base name for the output (without _output.xlsx). "
                         "Used with --json.")
    args = ap.parse_args()

    out_dir: Path = args.output
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.json:
        doc = parse_json_doc(args.json)
        stem = args.out_name or args.json.stem
        out_path = out_dir / f"{stem}_output.xlsx"
        write_output(doc, out_path)
        print(f"[ok] {args.json.name} -> {out_path.name} ({len(doc.lines)} lines)")
        return

    in_dir: Path = args.input
    if not in_dir.exists():
        print(f"Input dir missing: {in_dir}")
        sys.exit(1)

    if args.file:
        paths = [in_dir / args.file]
    else:
        paths = sorted(
            p for p in in_dir.iterdir()
            if p.suffix.lower() in (".xlsx", ".xls", ".html", ".pdf", *IMAGE_EXTS)
            and not p.name.startswith("~$")
        )

    if not paths:
        print("No processable files found.")
        return

    converted = 0
    vision_needed: list[str] = []
    for p in paths:
        ext = p.suffix.lower()
        if ext in IMAGE_EXTS:
            vision_needed.append(p.name)
            print(f"[vision] {p.name} — run via Claude vision (see SKILL.md).")
            continue
        out_path = convert_one(p, out_dir)
        if out_path:
            converted += 1
        elif ext == ".pdf":
            # Detect scanned PDFs that produced no text.
            try:
                import pdfplumber as _pp
                with _pp.open(p) as _pdf:
                    text = (_pdf.pages[0].extract_text() or "").strip()
                if not text:
                    vision_needed.append(p.name)
            except Exception:
                pass

    print(f"\nDone. {converted}/{len(paths)} files converted to {out_dir}")
    if vision_needed:
        print(f"\n{len(vision_needed)} file(s) need Claude vision (see SKILL.md Step 2):")
        for name in vision_needed:
            print(f"  - {name}")


if __name__ == "__main__":
    main()
