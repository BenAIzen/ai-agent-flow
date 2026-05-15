"""
Purchase-order converter web service.

POST /api/convert     multipart upload of one or more PO files
                      -> JSON: { summary: [...], download_id: "..." }
GET  /api/download/{id}  -> combined xlsx with the 7 requested columns

Images (.jpg/.jpeg/.png) and scanned PDFs (no extractable text) are reported
as skipped because vision is not wired up in this deployment.
"""

from __future__ import annotations

import io
import sys
import tempfile
import uuid
from pathlib import Path
from typing import Any

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

import openpyxl
from openpyxl.utils import get_column_letter

# Make the existing parser module importable.
SCRIPT_DIR = Path(__file__).resolve().parent  # purchase_order/web
PARSERS_DIR = SCRIPT_DIR.parent / "scripts"
sys.path.insert(0, str(PARSERS_DIR))

from convert_purchase_orders import (  # noqa: E402
    EXCEL_PARSERS,
    IMAGE_EXTS,
    SKIP_FILES,
    OrderDoc,
    parse_generic_xlsx,
    parse_html_generic,
    parse_pdf,
)

from extractors import extract_metadata  # noqa: E402

OUT_COLUMNS = ["발주날짜", "발주업장", "품목코드", "품명", "수량", "단위", "단가"]

# In-memory store for generated workbooks. Keys are UUIDs returned to the
# frontend; the bytes are deleted once downloaded (single-use).
_DOWNLOADS: dict[str, tuple[bytes, str]] = {}


def _parse_uploaded(path: Path) -> tuple[OrderDoc | None, str]:
    """Return (doc, status). status in {ok, skipped_image, skipped_scanned,
    skipped_unsupported, skipped_reference, failed}."""
    if path.name in SKIP_FILES:
        return None, "skipped_reference"
    ext = path.suffix.lower()
    if ext in IMAGE_EXTS:
        return None, "skipped_image"
    try:
        if ext in (".xlsx", ".xls"):
            parser = EXCEL_PARSERS.get(path.name) or parse_generic_xlsx
            doc = parser(path)
        elif ext == ".html":
            doc = parse_html_generic(path)
        elif ext == ".pdf":
            doc = parse_pdf(path)
            # If pdfplumber returned None *and* there's no extractable text,
            # treat as a scanned PDF.
            if doc is None or not doc.lines:
                try:
                    import pdfplumber
                    with pdfplumber.open(path) as pdf:
                        txt = (pdf.pages[0].extract_text() or "").strip()
                    if not txt:
                        return None, "skipped_scanned"
                except Exception:
                    pass
        else:
            return None, "skipped_unsupported"
    except Exception as e:
        print(f"[parse error] {path.name}: {e}", file=sys.stderr)
        return None, "failed"

    if doc is None or not doc.lines:
        return None, "failed"
    return doc, "ok"


def _build_workbook(rows: list[dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판매일보"
    ws.append(OUT_COLUMNS)
    for r in rows:
        ws.append([r.get(col, "") for col in OUT_COLUMNS])
    for col in range(1, len(OUT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


app = FastAPI(title="Purchase Order Converter")


@app.post("/api/convert")
async def convert(files: list[UploadFile] = File(...)) -> JSONResponse:
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    summary: list[dict[str, Any]] = []
    combined_rows: list[dict[str, Any]] = []

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        for uf in files:
            if not uf.filename:
                continue
            tmp_path = tmp / uf.filename
            with open(tmp_path, "wb") as out:
                out.write(await uf.read())

            doc, status = _parse_uploaded(tmp_path)
            entry: dict[str, Any] = {
                "file": uf.filename,
                "status": status,
                "date": "",
                "business": "",
                "line_count": 0,
                "lines": [],
            }

            if status != "ok" or doc is None:
                summary.append(entry)
                continue

            date, business = extract_metadata(tmp_path)
            entry["date"] = date
            entry["business"] = business
            entry["line_count"] = len(doc.lines)
            entry["lines"] = [
                {
                    "품목코드": line.품목코드,
                    "품명": line.품목명,
                    "수량": line.수량,
                    "단위": line.단위,
                    "단가": line.단가,
                }
                for line in doc.lines
            ]
            summary.append(entry)

            for line in doc.lines:
                combined_rows.append({
                    "발주날짜": date,
                    "발주업장": business,
                    "품목코드": line.품목코드,
                    "품명": line.품목명,
                    "수량": line.수량,
                    "단위": line.단위,
                    "단가": line.단가,
                })

    download_id = ""
    if combined_rows:
        download_id = uuid.uuid4().hex
        data = _build_workbook(combined_rows)
        _DOWNLOADS[download_id] = (data, f"발주서_변환_{download_id[:8]}.xlsx")

    return JSONResponse({
        "summary": summary,
        "download_id": download_id,
        "total_rows": len(combined_rows),
    })


@app.get("/api/download/{download_id}")
async def download(download_id: str) -> FileResponse:
    item = _DOWNLOADS.pop(download_id, None)
    if not item:
        raise HTTPException(status_code=404, detail="download expired or unknown")
    data, name = item
    # Write to a temp file so FileResponse can stream it; the OS will clean it
    # up shortly after the process closes the file.
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(data)
    tmp.flush()
    tmp.close()
    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=name,
    )


@app.get("/", response_class=HTMLResponse)
async def index() -> HTMLResponse:
    html_path = SCRIPT_DIR / "static" / "index.html"
    return HTMLResponse(html_path.read_text(encoding="utf-8"))


app.mount("/static", StaticFiles(directory=str(SCRIPT_DIR / "static")), name="static")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="127.0.0.1", port=8000, reload=False)
