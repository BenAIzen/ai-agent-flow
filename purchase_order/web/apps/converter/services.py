"""PO file parsing + result xlsx building.

Pure functions — no DB, no Django. Imported by views.py.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# Make the legacy parser package importable. It lives at
# purchase_order/scripts/convert_purchase_orders.py.
PARSERS_DIR = Path(__file__).resolve().parents[3] / "scripts"
if str(PARSERS_DIR) not in sys.path:
    sys.path.insert(0, str(PARSERS_DIR))

from convert_purchase_orders import (  # noqa: E402
    EXCEL_PARSERS,
    IMAGE_EXTS,
    SKIP_FILES,
    OrderDoc,
    parse_generic_xlsx,
    parse_html_generic,
    parse_pdf,
)

from .extractors import extract_metadata  # noqa: E402

OUT_COLUMNS = ["발주날짜", "발주업장", "품목코드", "품명", "수량", "단위", "단가"]


def parse_uploaded(path: Path) -> tuple[OrderDoc | None, str]:
    """Return (doc, status).

    status ∈ {ok, skipped_image, skipped_scanned, skipped_unsupported,
              skipped_reference, failed}.
    """
    if path.name in SKIP_FILES:
        return None, "skipped_reference"
    ext = path.suffix.lower()
    if ext in IMAGE_EXTS:
        return None, "skipped_image"
    try:
        if ext in (".xlsx", ".xls"):
            parser = EXCEL_PARSERS.get(path.name) or parse_generic_xlsx
            doc = parser(path)
        elif ext == ".html":
            doc = parse_html_generic(path)
        elif ext == ".pdf":
            doc = parse_pdf(path)
            if doc is None or not doc.lines:
                try:
                    import pdfplumber
                    with pdfplumber.open(path) as pdf:
                        txt = (pdf.pages[0].extract_text() or "").strip()
                    if not txt:
                        return None, "skipped_scanned"
                except Exception:
                    pass
        else:
            return None, "skipped_unsupported"
    except Exception as e:  # noqa: BLE001
        print(f"[parse error] {path.name}: {e}", file=sys.stderr)
        return None, "failed"

    if doc is None or not doc.lines:
        return None, "failed"
    return doc, "ok"


def build_workbook(rows: list[dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판매일보"
    ws.append(OUT_COLUMNS)
    for r in rows:
        ws.append([r.get(col, "") for col in OUT_COLUMNS])
    for col in range(1, len(OUT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Re-export for views.
__all__ = ["OUT_COLUMNS", "parse_uploaded", "build_workbook", "extract_metadata"]
