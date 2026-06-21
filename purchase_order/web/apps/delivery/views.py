import io

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.companies.mixins import CompanyScopedMixin, get_request_company
from apps.prices.models import lookup_sale_price

from .models import DeliveryOrder
from .serializers import DeliveryOrderSerializer


class DeliveryOrderViewSet(CompanyScopedMixin, viewsets.ModelViewSet):
    serializer_class = DeliveryOrderSerializer
    queryset = (
        DeliveryOrder.objects
        .select_related("partner")
        .prefetch_related("lines", "lines__item")
        .all()
    )

    def get_queryset(self):
        qs = super().get_queryset()
        date_from = self.request.query_params.get("from")
        date_to = self.request.query_params.get("to")
        partner = self.request.query_params.get("partner")
        q = self.request.query_params.get("q")
        if date_from:
            qs = qs.filter(order_date__gte=date_from)
        if date_to:
            qs = qs.filter(order_date__lte=date_to)
        if partner:
            qs = qs.filter(partner_id=partner)
        if q:
            qs = qs.filter(Q(order_no__icontains=q) | Q(partner__name__icontains=q)
                           | Q(note__icontains=q))
        return qs

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """필터 조건과 동일한 출고전표를 xlsx로 다운로드.

        한 라인 = 한 행 (master-detail flatten). 상단에 회사·기간 헤더 표시.
        """
        company = get_request_company(request)
        qs = self.get_queryset().order_by("order_date", "order_no")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "출고처리"

        # 헤더 정보
        date_from = request.query_params.get("from") or ""
        date_to = request.query_params.get("to") or ""
        ws.append([f"{company.name} - 출고처리"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([f"기간: {date_from} ~ {date_to}"])
        ws.append([])

        # 컬럼 헤더
        headers = [
            "출고번호", "출고일자", "거래처코드", "거래처명",
            "품목코드", "품명", "규격", "단위",
            "수량", "단가", "공급가액", "부가세", "합계",
            "VAT", "전표상태", "적요",
        ]
        ws.append(headers)
        head_row = ws.max_row
        header_fill = PatternFill("solid", fgColor="E2E8F0")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=head_row, column=c)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 데이터 행 (한 라인 = 한 행)
        total_supply = total_vat = total_amount = 0
        for order in qs:
            for line in order.lines.all():
                ws.append([
                    order.order_no, str(order.order_date),
                    order.partner.code, order.partner.name,
                    line.item.code, line.item.name,
                    line.spec or "", line.unit or "",
                    float(line.qty), float(line.unit_price),
                    float(line.supply_amount), float(line.vat_amount),
                    float(line.total),
                    order.get_vat_type_display(),
                    order.get_status_display(),
                    order.note or "",
                ])
                total_supply += float(line.supply_amount)
                total_vat += float(line.vat_amount)
                total_amount += float(line.total)

        # 합계 행
        if total_amount:
            ws.append([])
            sum_row = ws.max_row + 1
            ws.cell(row=sum_row, column=10, value="합계").font = Font(bold=True)
            ws.cell(row=sum_row, column=11, value=total_supply).font = Font(bold=True)
            ws.cell(row=sum_row, column=12, value=total_vat).font = Font(bold=True)
            ws.cell(row=sum_row, column=13, value=total_amount).font = Font(bold=True)
            for c in (11, 12, 13):
                ws.cell(row=sum_row, column=c).number_format = "#,##0"

        # 컬럼 너비 + 숫자 포맷
        widths = [16, 11, 8, 22, 10, 22, 12, 6, 8, 10, 12, 10, 12, 9, 8, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=head_row + 1, min_col=9, max_col=13):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.##" if cell.column == 9 else "#,##0"

        # Freeze 헤더
        ws.freeze_panes = f"A{head_row + 1}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        period = f"{date_from}_{date_to}" if date_from and date_to else "all"
        filename = f"출고처리_{period}.xlsx"

        from urllib.parse import quote
        # 한글 파일명은 RFC 5987 형식으로 안전 인코딩
        resp = HttpResponse(
            buf.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        resp["Content-Disposition"] = (
            f"attachment; filename*=UTF-8''{quote(filename)}"
        )
        return resp

    @action(detail=False, methods=["get"], url_path="suggest-price")
    def suggest_price(self, request):
        """출고처리 UI에서 거래처/품목 선택 시 단가 자동조회."""
        # 회사 컨텍스트 검증만 수행 (X-Company-Id 누락 시 403).
        get_request_company(request)
        try:
            partner_id = int(request.query_params.get("partner"))
            item_id = int(request.query_params.get("item"))
        except (TypeError, ValueError):
            return Response({"detail": "partner, item required"}, status=400)
        ymd = request.query_params.get("date")
        if not ymd:
            return Response({"detail": "date required"}, status=400)

        p = lookup_sale_price(partner_id, item_id, ymd)
        if p:
            return Response({
                "unit_price": p.sale_price,
                "source": "partner_price",
                "effective_from": p.effective_from,
            })

        # PartnerPrice 미등록 → 사용자가 수동 입력하도록 0 반환.
        # (표준원가 폴백 제거: 그린푸드는 거래처별 단가만 사용)
        return Response({"unit_price": 0, "source": "none"})
