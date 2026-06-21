"""xlsx에서 거래처 + 계좌 가져오기.

Usage:
    python manage.py import_partners --company "(주)그린푸드" \\
        --file ../../2.그린_매입거래처_계좌목록.xlsx

xlsx 시트 "거래처입금리스트_2023" 헤더:
    No. | 더존코드 | 업체명 | 예금주 | 은행 | 계좌번호 | 금액 | 입금메모 | 출금통장메모 | 비고
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.companies.models import Company
from apps.partners.models import Partner, PartnerAccount


SHEET = "거래처입금리스트_2023"
HEADER_ROW = 3   # 1-indexed


class Command(BaseCommand):
    help = "거래처 xlsx → DB 일괄 import"

    def add_arguments(self, parser):
        parser.add_argument("--company", required=True,
                            help="대상 회사명 (정확히 일치해야 함)")
        parser.add_argument("--file", required=True,
                            help="xlsx 파일 경로")
        parser.add_argument("--dry-run", action="store_true",
                            help="실제 저장 없이 시뮬레이션만")

    def handle(self, *args, **opts):
        company_name = opts["company"]
        try:
            company = Company.objects.get(name=company_name)
        except Company.DoesNotExist:
            raise CommandError(f"회사 '{company_name}' 을 찾을 수 없습니다")

        path = Path(opts["file"]).resolve()
        if not path.exists():
            raise CommandError(f"파일 없음: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)
        if SHEET not in wb.sheetnames:
            raise CommandError(
                f"시트 '{SHEET}' 가 없습니다. 가능한 시트: {wb.sheetnames}"
            )
        ws = wb[SHEET]

        created_partners = 0
        updated_partners = 0
        created_accounts = 0
        skipped_blank = 0

        for idx, r in enumerate(
            ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=1
        ):
            no, code, name, holder, bank, account_no, *_ = r + (None,) * 10

            if not name:
                skipped_blank += 1
                continue

            # 더존코드가 비어있으면 행번호 기반 코드 부여 (충돌 방지)
            partner_code = str(code).strip() if code else f"NOCODE-{idx:03d}"
            partner, p_created = Partner.objects.update_or_create(
                company=company, code=partner_code,
                defaults={
                    "name": str(name).strip(),
                    "biz_class": "vendor",  # xlsx는 매입처(지급)용
                    "rep_name": (str(holder).strip() if holder else ""),
                    "is_active": True,
                },
            )
            if p_created:
                created_partners += 1
            else:
                updated_partners += 1

            if bank and account_no:
                _, a_created = PartnerAccount.objects.get_or_create(
                    partner=partner,
                    bank=str(bank).strip(),
                    account_no=str(account_no).strip(),
                    defaults={
                        "holder": str(holder).strip() if holder else "",
                        "is_default": True,
                        "is_active": True,
                    },
                )
                if a_created:
                    created_accounts += 1

        if opts["dry_run"]:
            self.stdout.write(self.style.WARNING("DRY-RUN — DB 변경 안 함"))

        self.stdout.write(self.style.SUCCESS(
            f"+ partners  created={created_partners}  updated={updated_partners}"
        ))
        self.stdout.write(self.style.SUCCESS(
            f"+ accounts  created={created_accounts}"
        ))
        self.stdout.write(f"  (skipped blank rows: {skipped_blank})")
