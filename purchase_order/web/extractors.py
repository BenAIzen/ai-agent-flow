"""
Pull 발주날짜 (order date) and 발주업장 (business name) from purchase-order
file *content* — never from filename. Returns ("", "") when nothing matches.

The CLI converter (scripts/convert_purchase_orders.py) falls back to filename
and today's date; the web flow requires real document values, so this module
re-reads the source file independently of the parsers.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


# ---------------------------------------------------------------------------
# Date extraction
# ---------------------------------------------------------------------------

_MONTH_EN = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Labels in priority order. Each label maps to a regex that captures the date string.
# We try Korean 발주일자 first (true order date), then PO Date / 일자 variants.
_DATE_LABELS: list[tuple[str, str]] = [
    # 발주일자  2026-04-06  /  발주 일자: 2026/04/06
    ("발주일자", r"발\s*주\s*일\s*자\s*[:：]?\s*(\d{4}[-/.\s]\d{1,2}[-/.\s]\d{1,2})"),
    ("PO DATE",  r"PO\s*DATE\s*[:：]?\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})"),
    ("PO Date",  r"PO\s*Date\s*[:：]?\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})"),
    ("PO Submit Date", r"PO\s*Submit\s*Date\s*[:：]?\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})"),
    ("PO 제출 날짜", r"PO\s*제출\s*날짜\s*[:：]?\s*(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})"),
    ("발주 날짜", r"발\s*주\s*날\s*짜\s*[:：]?\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})"),
    # 일자 2026/04/06  — holisongdo style (avoid '입고일자' / '사용예정일')
    ("일자",     r"(?<![고예정])일\s*자\s*[:：]?\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})"),
    # IHG style: Delivery Date: 07 Apr 2026  (last-resort, since this is delivery not order)
    ("Delivery Date en-word", r"Delivery\s*Date\s*[:：]?\s*(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})"),
    ("Delivery date iso", r"Delivery\s*date\s*[:：]?\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})"),
    ("배송 날짜",  r"배\s*송\s*날\s*짜\s*[:：]?\s*(\d{1,2}[-/.]\d{1,2}[-/.]\d{4})"),
    # Plain "Delivery : 2026/03/16" (gangnam-fourpoints) — lowest priority
    ("Delivery iso", r"Delivery\s*[:：]\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})"),
]


def _to_iso(raw: str) -> str:
    """Normalize captured date string to YYYY-MM-DD. Empty string on failure."""
    s = raw.strip()
    # 07 Apr 2026  /  7 April 2026
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$", s)
    if m:
        d, mon_name, y = m.groups()
        mon = _MONTH_EN.get(mon_name.lower()[:4]) or _MONTH_EN.get(mon_name.lower()[:3])
        if mon:
            return f"{int(y):04d}-{mon:02d}-{int(d):02d}"
        return ""
    # MM/DD/YYYY or DD/MM/YYYY — heuristic: if first number > 12, it's DD/MM
    parts = re.split(r"[-/.\s]+", s)
    parts = [p for p in parts if p]
    if len(parts) != 3:
        return ""
    a, b, c = parts
    try:
        ai, bi, ci = int(a), int(b), int(c)
    except ValueError:
        return ""
    # YYYY first
    if ai >= 1900:
        y, mo, d = ai, bi, ci
    # YYYY last — disambiguate MM/DD vs DD/MM
    elif ci >= 1900:
        if ai > 12:
            d, mo, y = ai, bi, ci
        else:
            # default to MM/DD/YYYY (the Birchstreet-KR pages use it)
            mo, d, y = ai, bi, ci
    else:
        return ""
    if not (1 <= mo <= 12 and 1 <= d <= 31):
        return ""
    return f"{y:04d}-{mo:02d}-{d:02d}"


def extract_date_from_text(text: str) -> str:
    if not text:
        return ""
    for _, pat in _DATE_LABELS:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            iso = _to_iso(m.group(1))
            if iso:
                return iso
    return ""


# ---------------------------------------------------------------------------
# Business (발주업장) extraction
# ---------------------------------------------------------------------------

# Lines we should never return as a business name.
_BUSINESS_NEGATIVE = re.compile(
    r"^("
    r"PURCHASE\s*ORDER|구매\s*주문|발\s*주\s*서|RECEIVING\s*RECORD|거\s*래\s*명\s*세\s*표"
    r"|Order\s*For|Order\s*Sent|Order\s*By|Delivery|PO\s*N|PR\s*N|SUPPLIER|Vendor|VENDOR"
    r"|업체명|업\s*장|Account\s*Code|Cost\s*Centre"
    r"|GREEN\s*FOOD|그린푸드|\(주\)그린푸드|㈜그린푸드"
    r"|T\s*E\s*L|TEL|FAX|서울|인천|경기"
    r"|주\s*소|ADDRESS|Phone|Email|이메일|전화|팩스|Telephone"
    r")",
    flags=re.IGNORECASE,
)


def _looks_like_business(line: str) -> bool:
    s = line.strip()
    if not s or len(s) < 2 or len(s) > 80:
        return False
    if s in _HEADER_WORDS:
        return False
    if _BUSINESS_NEGATIVE.match(s):
        return False
    # Skip pure number/code lines
    if re.match(r"^[\d\-\.\s/:]+$", s):
        return False
    # Skip lines that are just labels
    if s.endswith(":") or s.endswith("："):
        return False
    # Skip address-like lines that start with a number (e.g., "52 Toegye-ro")
    if re.match(r"^\d+\s", s):
        return False
    return True


_LABEL_BUSINESS_PATTERNS = [
    # Korean 발주서 top line: "발 주 서  ROKAUS HOTEL"
    r"발\s*주\s*서\s+([^\n\r]{2,80})",
    # Birchstreet-style DELIVERY TO with optional trailing ADDRESS
    r"DELIVERY\s*TO\s*[:：]\s*([^\n\r]+?)(?=\s+ADDRESS|\s+CONTACT|\s+PO\s|\s+TEL|\s*[\n\r]|$)",
    r"BILL\s*TO\s*[:：]\s*([^\n\r]+?)(?=\s+ADDRESS|\s+CONTACT|\s+TEL|\s*[\n\r]|$)",
    r"배\s*송\s*[:：]\s*([^\n\r]+?)(?=\s+\d|\s+서울|\s+인천|\s*[\n\r]|$)",
    # 거래명세표 layout (홀리송도): "받 ... 상 호 <buyer>" — buyer is the customer.
    # Anchor on the "받" side to avoid grabbing the vendor (공급자) row.
    r"받[\s\S]{0,40}?상\s*호\s+([^\n\r]+?)(?=\s+대\s+|\s{2,}|\s+표|[\n\r]|$)",
]


_HEADER_WORDS = {
    "No", "no", "NO", "품목명", "품명", "재고규격", "구매규격", "규격",
    "수량", "단위", "단가", "금액", "비고", "날짜", "일자", "그린푸드 발주서",
    "발주서", "품목코드", "Item Code", "Item Name", "Description",
}


def extract_business_from_text(text: str) -> str:
    if not text:
        return ""
    # First try labeled patterns
    for pat in _LABEL_BUSINESS_PATTERNS:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            cand = m.group(1).strip().rstrip(",;")
            if _looks_like_business(cand) and not _is_vendor_us(cand):
                return cand
    # Fallback: scan first ~12 non-empty lines, pick first business-looking line
    # that isn't us (Green Food / 그린푸드) or a label.
    for raw in text.splitlines()[:20]:
        line = raw.strip()
        if not line:
            continue
        if _is_vendor_us(line):
            continue
        if _looks_like_business(line):
            return line
    return ""


def _is_vendor_us(s: str) -> bool:
    """Lines that name us (the supplier), not the customer."""
    s_low = s.lower()
    return any(k in s_low for k in ("green food", "greenfood")) or any(
        k in s for k in ("그린푸드", "(주)그린푸드", "㈜그린푸드")
    )


# ---------------------------------------------------------------------------
# Format dispatch
# ---------------------------------------------------------------------------

def extract_from_xlsx(path: Path) -> tuple[str, str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    # Aggregate the first 8 rows of every sheet into a search corpus,
    # but also keep cells positionally for the cell-label scan.
    found_date = ""
    found_biz = ""

    for sn in wb.sheetnames:
        ws = wb[sn]
        # Build text + cell grid up to row 10.
        rows: list[list[str]] = []
        for r in range(1, min(ws.max_row, 12) + 1):
            row = []
            for c in ws[r]:
                v = c.value
                if v is None:
                    row.append("")
                else:
                    row.append(str(v))
            rows.append(row)
        text = "\n".join(" ".join(r) for r in rows)

        if not found_date:
            found_date = extract_date_from_text(text)
        if not found_date:
            # Cell-level: pick up any cell value that itself parses as an ISO date
            for row in rows:
                for v in row:
                    iso = _cell_to_iso(v)
                    if iso:
                        found_date = iso
                        break
                if found_date:
                    break

        if not found_biz:
            # Label-cell scan: cell containing 업체명 / 업장 — value is in the
            # right-neighboring non-empty cell on the same row.
            for row in rows:
                for i, v in enumerate(row):
                    label = v.strip() if isinstance(v, str) else ""
                    if label in ("업체명", "업장", "거래처명", "업장명"):
                        for j in range(i + 1, len(row)):
                            cand = (row[j] or "").strip()
                            if cand and not _is_vendor_us(cand):
                                found_biz = cand
                                break
                        if found_biz:
                            break
                if found_biz:
                    break
        if not found_biz:
            # Detect "row 0 is the table header" case: if 3+ cells in row 0 are
            # well-known column-header words, this sheet has no title row, so
            # there is no business name to find here.
            header_like = {"품목명", "품명", "품목코드", "규격", "구매규격", "재고규격",
                           "단위", "품목단위", "재고단위", "수량", "단가", "금액", "비고",
                           "Item Code", "Item Name", "Description", "Qty", "Unit Price"}
            row0 = rows[0] if rows else []
            hits = sum(1 for v in row0 if isinstance(v, str) and v.strip() in header_like)
            if hits < 3:
                # Anything that also appears on row 2-4 is a column header, not a title.
                header_set: set[str] = set()
                for hr in rows[1:5]:
                    for v in hr:
                        s = (v or "").strip() if isinstance(v, str) else ""
                        if s:
                            header_set.add(s)
                for v in row0:
                    s = (v or "").strip() if isinstance(v, str) else ""
                    if not s or s in header_set:
                        continue
                    if _is_vendor_us(s):
                        continue
                    if not _looks_like_business(s):
                        continue
                    if "팩스" in s or "TEL" in s.upper() or "FAX" in s.upper():
                        continue
                    if re.match(r"^\d{4}", s):
                        continue
                    # Reject obvious title rows like "그린푸드 6월 18일"
                    if re.search(r"\d{1,2}\s*월\s*\d{1,2}\s*일", s):
                        continue
                    found_biz = s
                    break

        if found_date and found_biz:
            break

    return found_date, found_biz


def _cell_to_iso(v: str) -> str:
    if not v:
        return ""
    s = str(v).strip()
    # 2026-04-07 or 2026/04/07 (optionally with trailing time)
    m = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", s)
    if m:
        y, mo, d = (int(x) for x in m.groups())
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return ""


def extract_from_xls(path: Path) -> tuple[str, str]:
    if xlrd is None:
        return "", ""
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    rows: list[list[str]] = []
    for r in range(min(ws.nrows, 12)):
        rows.append([str(v) if v is not None else "" for v in ws.row_values(r)])
    text = "\n".join(" ".join(r) for r in rows)
    found_date = extract_date_from_text(text)
    if not found_date:
        for row in rows:
            for v in row:
                iso = _cell_to_iso(v)
                if iso:
                    found_date = iso
                    break
            if found_date:
                break
    # hoam-style: a cell contains "영 업 장 :    022   직원식당(2)" all in one cell.
    found_biz = ""
    for r in range(min(ws.nrows, 30)):
        for v in ws.row_values(r):
            s = str(v) if v is not None else ""
            m = re.match(r"^\s*영\s*업\s*장\s*[:：]\s*\d*\s+([^\d].+)$", s)
            if m:
                cand = m.group(1).strip()
                if _looks_like_business(cand) and not _is_vendor_us(cand):
                    found_biz = cand
                    break
        if found_biz:
            break
    if not found_biz:
        found_biz = extract_business_from_text(text)
    return found_date, found_biz


def extract_from_html(path: Path) -> tuple[str, str]:
    if BeautifulSoup is None:
        return "", ""
    with open(path, "rb") as f:
        soup = BeautifulSoup(f.read(), "lxml")
    text = soup.get_text("\n", strip=True)
    found_date = extract_date_from_text(text)
    found_biz = ""

    # Birchstreet pages often have the hotel name directly under "PURCHASE ORDER".
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for i, ln in enumerate(lines):
        if re.match(r"^(PURCHASE\s*ORDER|구매\s*주문|발\s*주\s*서)", ln, flags=re.IGNORECASE):
            for nxt in lines[i + 1: i + 5]:
                if _looks_like_business(nxt) and not _is_vendor_us(nxt):
                    found_biz = nxt
                    break
            if found_biz:
                break
    if not found_biz:
        found_biz = extract_business_from_text(text)
    return found_date, found_biz


def extract_from_pdf(path: Path) -> tuple[str, str]:
    if pdfplumber is None:
        return "", ""
    try:
        with pdfplumber.open(path) as pdf:
            text = pdf.pages[0].extract_text() or ""
    except Exception:
        return "", ""

    found_date = extract_date_from_text(text)
    found_biz = ""

    # Look for known label first
    for pat in _LABEL_BUSINESS_PATTERNS:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            cand = m.group(1).strip().rstrip(",;")
            if _looks_like_business(cand) and not _is_vendor_us(cand):
                found_biz = cand
                break

    if not found_biz:
        # The hotel name is typically on line 1 or 2 of the PDF.
        for ln in text.splitlines()[:6]:
            s = ln.strip()
            if not s:
                continue
            if _is_vendor_us(s):
                continue
            if _looks_like_business(s) and "PURCHASE ORDER" not in s.upper():
                found_biz = s
                break

    return found_date, found_biz


def extract_metadata(path: Path) -> tuple[str, str]:
    """Dispatch by extension. Returns (발주날짜, 발주업장), each "" if not found."""
    ext = path.suffix.lower()
    try:
        if ext == ".xlsx":
            return extract_from_xlsx(path)
        if ext == ".xls":
            return extract_from_xls(path)
        if ext == ".html":
            return extract_from_html(path)
        if ext == ".pdf":
            return extract_from_pdf(path)
    except Exception:
        return "", ""
    return "", ""
